# -*- coding: utf-8 -*-
"""
Created on Sun Dec  5 17:16:18 2021

@author: Golden Snow
"""

import numpy as np
import openpyxl
import webbrowser

def curvefitting(moneyness, volatility, picTitle, vol, n):
    print(f"moneyness: {moneyness}")
    print(f"volatility: {volatility}")
    print(f"picTitle: {picTitle}")
    print(f"vol: {vol}")
    print(n)
    
    wb = openpyxl.load_workbook('Output4.0.xlsx')
    ws = wb['1M'] 
    
    for m in range(1,n+1):
        maxValue = max(volatility[m])  
        c = np.polyfit(moneyness, volatility[m], deg=2)
        p = np.poly1d(c)
        print(f"p: {p}")
        
        spread = p(1) - vol[m-1]
        print(f"p(1): {p(1)}")
        print(f"spread: {spread}")

        p1 = p - spread
        print(f"p1: {p1}")
     
        for i in range(0,11):
            ws.cell(row=i+5, column=m+1).value = round(p1(moneyness[i]),2)
        
    wb.save('Output4.0.xlsx')
    webbrowser.open('Output4.0.xlsx')
       
    return [c[0], c[1], maxValue]
       


if __name__ == '__main__':

    wb = openpyxl.load_workbook("Input4.0.xlsx")
    ws = wb['1M']
    x = []
    y = [[]]
    
    # 设置场外期权目标波动率
    vol = [41.99,38.87,63.36,14.27,27.93,15.9,29.05,10.12,16.42,19.76,45.98,22.23,33.54,17.43,30.02,11.31,20.24,28.62,24.85,17.39,22.18,43.99,45.21,32.13,12.98,20.76,31.88,33.6,29.56,21.28,22.39,24.41,36.6,54.08,73.69,47.4,58.62]
  
    # 场内期权对应编号 例:铁矿石对应编号001 即n=1   
    for col in ws.iter_cols(min_row=5,max_row=15,min_col=1,max_col=1):
        for cell in col:
            x.append(cell.value)
            
    for n in range(1,38):  
        y.append([])          
        for col in ws.iter_cols(min_row=5,max_row=15,min_col=n+1,max_col=n+1):
            for cell in col:
                y[n].append(cell.value)
                
print('================================')
print(x)
print('================================')
print(y)

curvefitting(x, y, "Volatility VS Moneyness", vol, n)

